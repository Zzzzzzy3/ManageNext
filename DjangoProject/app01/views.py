from django.http import HttpResponse
from django.shortcuts import render, redirect


from django.http import JsonResponse

def login(request):
    return render(request, 'login.html')
def user_login(request):
    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        try:
            user = User.objects.get(name=username, password=password)
            return JsonResponse({'status': 'success'})  # 简化响应格式
        except User.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '用户名或密码错误'}, status=400)
    
    return JsonResponse({'status': 'error', 'message': '无效请求'}, status=405)
def login_register(request):
    if request.method == 'POST':
        # 从POST请求中获取参数
        print("fuck")
        username = request.POST.get('reg_username')
        password = request.POST.get('reg_password')
        email = request.POST.get('email')

        # 参数校验

        # 检查用户名是否存在（使用模型中的name字段）
        if User.objects.filter(name=username).exists():
            return JsonResponse({'success': False, 'message': '用户名已存在'}, status=400)

        try:
            # 创建用户
            User.objects.create(
                name=username,
                email=email,
                password=password

            )
            return JsonResponse({'success': True})
            
        except Exception as e:
            # 记录错误日志
            print(f"注册失败: {str(e)}")
            return JsonResponse({'success': False, 'message': '服务器内部错误'}, status=500)

    return JsonResponse({'success': False, 'message': '非法请求方法'}, status=405)


from django.http import JsonResponse
from urllib.parse import parse_qs, unquote  # 添加unquote解码
import time
from django.utils import timezone
from django.views.decorators.http import require_GET
from pyexpat.errors import messages
from django.views.decorators.csrf import csrf_exempt
from app01.models import DishTable
from app01.models import Supplier
from django.db import IntegrityError
from .models import User, Inventory
from .manage_function import func_sale
import openpyxl
import requests
# Create your views here.

def manage(request):
    user_login= User.objects.all()
    dish_table = DishTable.objects.all()
    supplier = Supplier.objects.all()
    inventory = Inventory.objects.all()
    dish_sales,inventory_number=func_sale(dish_table, inventory)

    return render(request, "main_managenext.html", { 'user_login':user_login,'dish_table': dish_table,'supplier': supplier, 'inventory': inventory, 'dish_sales': dish_sales, 'inventory_number': inventory_number })
#用户模块
def user_delete(request):
    if request.method == "POST":
        user_id = request.POST.get("user_id")
        try:
            user_row = User.objects.get(id=user_id)
            user_row.delete()
            return JsonResponse({"status": "success", "message": "删除成功"})
        except DishTable.DoesNotExist:
            return JsonResponse({"status": "error", "message": "不存在"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error"}, status=405)
#菜品
def dish_table_add(request):
    if request.method == "POST":
        dish_name = request.POST.get("dish_name")
        dish_amount = request.POST.get("dish_amount")
        dish_price = request.POST.get("dish_price")

        if not dish_name:
            return JsonResponse({"status": "error", "message": "菜品名称不能为空"}, status=400)

        if DishTable.objects.filter(dish_name=dish_name).exists():
            return JsonResponse({"status": "error", "message": "菜品名称已存在"}, status=400)
        try:
            # 创建菜单并保存到数据库 函数名是models模型
            dish = DishTable(dish_name=dish_name, dish_amount=dish_amount,dish_price=dish_price)
            dish.save()
            print("菜品添加成功")
            return JsonResponse({
                "status": "success",
                "message": "菜品添加成功",
                "dish": {
                    "id": dish.id,
                    "name": dish_name,
                    "amount": dish_amount,
                    "price": dish_price,
                }
            })
        except IntegrityError:
            # 处理菜品名重复的情况
            return JsonResponse(
                {"status": "error", "message": "菜品已存在"},
                status=400
            )

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)


    return JsonResponse({"status": "error"}, status=405)

def dish_table_delete(request):
    if request.method == "POST":
        dish_id = request.POST.get("dish_id")
        try:
            dish = DishTable.objects.get(id=dish_id)
            dish.delete()
            return JsonResponse({"status": "success", "message": "菜品删除成功"})
        except DishTable.DoesNotExist:
            return JsonResponse({"status": "error", "message": "菜品不存在"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error"}, status=405)

def dish_table_edit(request):
    if request.method == "PUT":
        try:
            # 参数解析
            body_data = request.body.decode('utf-8')
            params = parse_qs(body_data)

            # 解码参数
            dish_id = params.get('id', [''])[0]
            dish = DishTable.objects.get(id=dish_id)

            # 对中文参数进行解码
            dish.dish_name = unquote(params.get('dish_name', [''])[0])  # 解码中文
            dish.dish_amount = int(params.get('dish_amount', [0])[0])
            dish.dish_price = float(params.get('dish_price', [0])[0])
            dish.save()

            return JsonResponse({
                "status": "success",
                "message": "更新成功",
                "dish": {
                    "id": dish.id,
                    "dish_name": dish.dish_name,  # 显示中文
                    "dish_amount": dish.dish_amount,
                    "dish_price": str(dish.dish_price)
                }
            })
        # except DishTable.DoesNotExist:
        #     return JsonResponse({"status": "error", "message": "菜品不存在"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error"}, status=405)

@require_GET
def dish_table_query(request):
    try:
        # 获取查询参数
        dish_id = request.GET.get('id')
        dish_name = request.GET.get('name', '').strip()

        # 构建查询条件
        query_params = {}
        if dish_id and dish_id.isdigit():
            query_params['id'] = dish_id
        if dish_name:
            query_params['dish_name__icontains'] = dish_name

        # 执行查询
        dishes = DishTable.objects.filter(**query_params).values(
            'id',
            'dish_name',
            'dish_amount',
            'dish_price'
        )

        return JsonResponse({
            "status": "success",
            "count": dishes.count(),
            "data": list(dishes)
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f"查询失败: {str(e)}"
        }, status=500)
#供应商
def supplier_add(request):
    if request.method == "POST":
        name = request.POST.get("name")
        contact_person = request.POST.get("contact_person")
        phone=request.POST.get("phone")
        email=request.POST.get("email")
        address = request.POST.get("address")

        if not name:
            return JsonResponse({"status": "error", "message": "名称不能为空"}, status=400)

        if Supplier.objects.filter(name=name).exists():
            return JsonResponse({"status": "error", "message": "供应商名称已存在"}, status=400)
        try:
            # 创建菜单并保存到数据库 函数名是models模型
            supplier = Supplier(name=name, contact_person=contact_person, phone=phone, email=email, address=address)
            supplier.save()
            updated_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())

            print("供应商添加成功")
            return JsonResponse({
                "status": "success",
                "message": "供应商添加成功",
                "supplier": {
                    "id": supplier.supplier_id,
                    "name": name,
                    "contact_person": contact_person,
                    "phone": phone,
                    "email": email,
                    "address": address,
                    "updated_time": updated_time,
                }
            })
        except IntegrityError:
            # 处理名重复的情况
            return JsonResponse(
                {"status": "error", "message": "名称已存在"},
                status=400
            )

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error"}, status=405)

def supplier_delete(request):
    if request.method == "POST":
        supplier_id = request.POST.get("supplier_id")
        try:
            print("aas")
            supplier = Supplier.objects.get(supplier_id=supplier_id)

            supplier.delete()
            return JsonResponse({"status": "success", "message": "删除成功"})
        except Supplier.DoesNotExist:
            return JsonResponse({"status": "error", "message": "名称不存在"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error"}, status=405)

def supplier_edit(request):
    if request.method == "PUT":
        try:
            # 参数解析
            body_data = request.body.decode('utf-8')
            params = parse_qs(body_data)

            # 解码
            supplier_id = params.get('id', [''])[0]
            supplier = Supplier.objects.get(supplier_id=supplier_id)
            print(supplier)
            # 对中文参数进行解码
            supplier.name = unquote(params.get('name', [''])[0])  # 解码中文
            supplier.contact_person = unquote(params.get('contact_person', [''])[0])
            supplier.phone = unquote(params.get('phone', [''])[0])
            supplier.email = unquote(params.get('email', [''])[0])
            supplier.address = unquote(params.get('address', [''])[0])
            supplier.updated_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())

            supplier.save()

            return JsonResponse({
                "status": "success",
                "message": "更新成功",
                "supplier": {
                    #下面这个id传入给data.supplier.id
                    'id': supplier_id,
                    'name': supplier.name,
                    'contact_person': supplier.contact_person,
                    'phone': supplier.phone,
                    'email': supplier.email,
                    'address': supplier.address,
                    'updated_time': supplier.updated_time,
                }
            })
        # except DishTable.DoesNotExist:
        #     return JsonResponse({"status": "error", "message": "不存在"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error"}, status=405)

@require_GET
def supplier_query(request):
    try:
        # 获取查询参数
        supplier_id = request.GET.get('supplier_id')
        supplier_name = request.GET.get('name', '').strip()

        # 构建查询条件
        query_params = {}
        if supplier_id and supplier_id.isdigit():
            query_params['supplier_id'] = supplier_id
        if supplier_name:
            query_params['name__icontains'] = supplier_name

        # 执行查询
        suppliers = Supplier.objects.filter(**query_params).values(
            'supplier_id',
            'name',
            'contact_person',
            'phone',
            'email',
            'address',
            'updated_time',
        )

        return JsonResponse({
            "status": "success",
            "count": suppliers.count(),
            "data": list(suppliers)
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f"查询失败: {str(e)}"
        }, status=500)
#库存
def inventory_add(request):
    if request.method == "POST":
        product_name = request.POST.get("product_name")
        amount = request.POST.get("amount")
        expiration_data = request.POST.get("expiration_data")
        dish = request.POST.get("dish")
        warehouse_loc=request.POST.get("warehouse_loc")
        batch_no = request.POST.get("batch_no")
        category=request.POST.get("category")

        if not product_name:
            return JsonResponse({"status": "error", "message": "名称不能为空"}, status=400)

        try:
            # 创建菜单并保存到数据库 函数名是models模型
            inventory = Inventory(product_name=product_name,amount=amount,expiration_data=expiration_data,dish_id=dish,warehouse_loc=warehouse_loc,batch_no= batch_no,category=category)
            inventory.save()
            print(inventory.inventory_id)
            print("库存添加成功")
            return JsonResponse({
                "status": "success",
                "message": "库存添加成功",
                "inventory": {
                    "id": inventory.inventory_id,
                    "product_name": product_name,
                    "amount": amount,
                    "expiration_data": expiration_data,
                    "dish": dish,
                    "warehouse_loc": warehouse_loc,
                    "batch_no": batch_no,
                    "category": category,
                }
            })
        except IntegrityError:
            # 处理名重复的情况
            return JsonResponse(
                {"status": "error", "message": "名称已存在"},
                status=400
            )

        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)

    return JsonResponse({"status": "error"}, status=405)

def inventory_delete(request):
    if request.method == "POST":
        inventory_id = request.POST.get("inventory_id")
        print(request.POST.get("inventory_id"))

        try:
            inventory = Inventory.objects.get(inventory_id=inventory_id)
            print(inventory)
            inventory.delete()
            return JsonResponse({"status": "success", "message": "删除成功"})
        except Supplier.DoesNotExist:
            return JsonResponse({"status": "error", "message": "名称不存在"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error"}, status=405)

def inventory_edit(request):
    if request.method == "PUT":
        try:
            # 参数解析
            body_data = request.body.decode('utf-8')
            params = parse_qs(body_data)

            # 解码参数
            inventory_id = params.get('id', [''])[0]
            inventory = Inventory.objects.get(inventory_id=inventory_id)

            # 对中文参数进行解码
            inventory.product_name = unquote(params.get('product_name', [''])[0])  # 解码中文
            inventory.amount = unquote(params.get('amount', [''])[0])
            inventory.expiration_data = unquote(params.get('expiration_data', [''])[0])
            inventory.dish_id =  unquote(params.get('dish', [''])[0])
            print(inventory.dish)
            inventory.warehouse_loc = unquote(params.get('warehouse_loc', [''])[0])
            inventory.batch_no = unquote(params.get('batch_no', [''])[0])
            inventory.category = unquote(params.get('category', [''])[0])

            inventory.save()

            return JsonResponse({
                "status": "success",
                "message": "更新成功",
                "inventory": {
                    #下面这个id传入给data.xxx.id
                    'id': inventory_id,
                    'product_name': inventory.product_name,
                    'amount': inventory.amount,
                    'expiration_data': inventory.expiration_data,
                    'dish': inventory.dish_id,
                    'warehouse_loc': inventory.warehouse_loc,
                    'batch_no': inventory.batch_no,
                    'category': inventory.category,

                }
            })
        # except DishTable.DoesNotExist:
        #     return JsonResponse({"status": "error", "message": "不存在"}, status=404)
        except Exception as e:
            return JsonResponse({"status": "error", "message": str(e)}, status=500)
    return JsonResponse({"status": "error"}, status=405)

@require_GET
def inventory_query(request):
    try:
        # 获取查询参数
        inventory_id = request.GET.get('inventory_id')
        inventory_name = request.GET.get('product_name', '').strip()

        # 构建查询条件
        query_params = {}
        if inventory_id and inventory_id.isdigit():
            query_params['inventory_id'] = inventory_id
        if inventory_name:
            query_params['name__icontains'] = inventory_name
        # 执行查询
        inventorys = Inventory.objects.filter(**query_params).values(
            'inventory_id',
            'product_name',
            'amount',
            'expiration_data',
            'dish',
            'warehouse_loc',
            'batch_no',
            'category',

        )

        return JsonResponse({
            "status": "success",
            "count": inventorys.count(),
            "data": list(inventorys)
        })

    except Exception as e:
        return JsonResponse({
            "status": "error",
            "message": f"查询失败: {str(e)}"
        }, status=500)
#导出模块
def export_to_excel_dishtable(request):
    # 创建 Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '导出数据'

    # 写入表头（字段）
    sheet.append(['菜品ID', '菜品名称', '菜品前台消费数量','菜品价格'])

    # 写入数据
    for item in DishTable.objects.all():
        sheet.append([item.id, item.dish_name, item.dish_amount,item.dish_price])

    # 创建 HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exported_data_dishtable.xlsx'

    # 保存 Excel 到响应
    workbook.save(response)

    return response

def export_to_excel_inventory(request):
    # 创建 Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = '导出数据'

    # 写入表头（字段）
    sheet.append(['库存记录唯一标识','库内商品名称','库内商品数量','保质期限','关联菜品ID','仓库位置','批次号','分类标签'])

    # 写入数据
    for item in Inventory.objects.all():
        sheet.append([item.inventory_id,item.product_name,item.amount,item.expiration_data,item.dish_id,item.warehouse_loc,item.batch_no,item.category])

    # 创建 HttpResponse
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exported_data_inventory.xlsx'

    # 保存 Excel 到响应
    workbook.save(response)

    return response

def export_to_excel_supplier(request):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title='导出数据'

    sheet.append(['供应商唯一标识','供应商名称','主要联系人姓名','联系电话','联系邮箱','详细地址','最后修改时间'])
    for item in Supplier.objects.all():
        sheet.append([item.supplier_id,item.name,item.contact_person,item.phone,item.email,item.address,item.updated_time])

    response=HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=exported_data_supplier.xlsx'
    workbook.save(response)
    return response
#图标模块
def chart_data(request):
    data = DishTable.objects.all()

    names = [p.dish_name for p in data]
    sales = [p.dish_amount for p in data]
    amount=[]
    j=0
    for item in data:
        for i in item.inventory.all():
           j=j+i.amount
        amount.append(j)
        j=0
    return JsonResponse({
        'names': names,
        'sales': sales,
        'amount': amount
    })

def chart_data2(request):
    data = DishTable.objects.all()
    names = [p.dish_name for p in data]
    sales = [p.dish_amount for p in data]
    chartlist = []
    for  name,value in zip(names, sales):
        chartlist.append({"value": value, "name": name})
    return JsonResponse({
       'chartlist': chartlist

    })
#ai
import json

def ai_chat(request):

    url = "https://api.siliconflow.cn/v1/chat/completions"
    if request.method == "POST":
        try:
            # 获取用户输入
            data = json.loads(request.body)
            user_message = data["prompt"]
            payload = {
                "model": "Qwen/QwQ-32B",
                "messages": [
                    {
                        "role": "user",
                        "content": user_message
                    }
                ],
                "stream": False,
                "max_tokens": 5000,
                "stop": None,
                "temperature": 0.7,
                "top_p": 0.7,
                "top_k": 50,
                "frequency_penalty": 0.5,
                "n": 1,
                "response_format": {"type": "text"},
            }
            headers = {
                "Authorization": "Bearer sk-flfvazdlopmbsvydifueadczsvobkfszlqydgjtkkrebiwfe",
                "Content-Type": "application/json"
            }
            response = requests.request("POST", url, json=payload, headers=headers)
            response_data = response.json()
            ai_response = response_data["choices"][0]["message"]["content"]
            return JsonResponse({'reply': ai_response})

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request'}, status=400)

